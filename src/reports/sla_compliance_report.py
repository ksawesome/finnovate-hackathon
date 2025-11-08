"""
SLA Compliance Report Generator

Generates SLA compliance tracking with at-risk items, breach detection,
and deadline monitoring across GL accounts.
"""

from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from src.db import get_postgres_session
from src.db.postgres import get_gl_accounts_by_entity_period, get_responsibility_assignments
from src.reports import BaseReport, register_report


@register_report("sla_compliance")
class SLAComplianceReport(BaseReport):
    """Generate SLA compliance and deadline tracking report."""

    def generate(self) -> dict[str, str]:
        """
        Generate SLA compliance report in PDF and Excel formats.

        Returns:
            Dict with 'pdf' and 'excel' keys mapping to file paths
        """
        self.generated_at = datetime.now()

        # Fetch data
        data = self._fetch_data()

        # Generate Excel (primary format with multi-sheet breakdown)
        excel_path = self._generate_excel(data)
        self.generated_files["excel"] = str(excel_path)

        # Generate PDF summary
        pdf_path = self._generate_pdf(data)
        self.generated_files["pdf"] = str(pdf_path)

        return self.generated_files

    def _fetch_data(self) -> dict:
        """Fetch SLA compliance data from GL accounts."""
        session = get_postgres_session()

        try:
            # Get all GL accounts
            accounts = get_gl_accounts_by_entity_period(session, self.entity, self.period)

            # Get responsibility assignments for SLA tracking
            assignments = get_responsibility_assignments(session, self.entity, self.period)

            # Build assignment map
            assignment_map = {a.gl_account_id: a for a in assignments}

            # Categorize accounts by SLA status
            sla_data = self._categorize_by_sla(accounts, assignment_map)

            return {
                "sla_data": sla_data,
                "accounts": accounts,
                "assignments": assignments,
                "entity": self.entity,
                "period": self.period,
                "generated_at": self.generated_at,
            }
        except Exception as e:
            return {
                "error": str(e),
                "entity": self.entity,
                "period": self.period,
                "generated_at": self.generated_at,
            }
        finally:
            session.close()

    def _categorize_by_sla(self, accounts: list, assignment_map: dict) -> dict:
        """Categorize accounts by SLA compliance status."""
        now = datetime.now()

        on_time = []
        delayed = []
        at_risk = []  # Within 24h of deadline
        breached = []
        critical_sla = []

        for account in accounts:
            assignment = assignment_map.get(account.id)

            # Get deadline (default: review_date + 7 days if not specified)
            if hasattr(account, "sla_deadline") and account.sla_deadline:
                deadline = account.sla_deadline
            elif hasattr(account, "review_date") and account.review_date:
                deadline = account.review_date + timedelta(days=7)
            else:
                deadline = now + timedelta(days=7)  # Default 7 days from now

            # Calculate time to deadline
            time_to_deadline = (deadline - now).total_seconds() / 3600  # Hours

            # Build account info
            account_info = {
                "account_code": account.account_code,
                "account_name": getattr(account, "account_name", ""),
                "review_status": getattr(account, "review_status", "pending"),
                "criticality": getattr(account, "criticality", "Medium"),
                "closing_balance": getattr(account, "closing_balance", 0),
                "deadline": deadline,
                "time_to_deadline_hours": time_to_deadline,
                "reviewer": (
                    assignment.reviewer.full_name
                    if assignment and assignment.reviewer
                    else "Unassigned"
                ),
                "flagged": getattr(account, "flagged", False),
            }

            # Categorize
            if account_info["review_status"] == "reviewed":
                on_time.append(account_info)
            elif time_to_deadline < 0:
                breached.append(account_info)
            elif time_to_deadline <= 24:
                at_risk.append(account_info)
            elif time_to_deadline > 168:  # More than 7 days
                on_time.append(account_info)
            else:
                delayed.append(account_info)

            # Track critical SLA items
            if account_info["criticality"] in ["High", "Critical"]:
                critical_sla.append(account_info)

        # Sort by time to deadline (ascending - most urgent first)
        at_risk.sort(key=lambda x: x["time_to_deadline_hours"])
        breached.sort(key=lambda x: x["time_to_deadline_hours"])
        critical_sla.sort(key=lambda x: x["time_to_deadline_hours"])

        # Calculate compliance metrics
        total_accounts = len(accounts)
        compliant_count = len(on_time)
        at_risk_count = len(at_risk)
        breached_count = len(breached)

        compliance_rate = (compliant_count / total_accounts * 100) if total_accounts > 0 else 0

        return {
            "on_time": on_time,
            "delayed": delayed,
            "at_risk": at_risk,
            "breached": breached,
            "critical_sla": critical_sla,
            "total_accounts": total_accounts,
            "compliant_count": compliant_count,
            "at_risk_count": at_risk_count,
            "breached_count": breached_count,
            "compliance_rate": compliance_rate,
        }

    def _generate_excel(self, data: dict) -> Path:
        """Generate multi-sheet Excel workbook."""
        output_path = self._get_output_path("xlsx")

        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        sla_data = data.get("sla_data", {})

        # Sheet 1: Summary
        self._create_summary_sheet(wb, data, sla_data)

        # Sheet 2: On-Time
        self._create_category_sheet(wb, "On-Time", sla_data.get("on_time", []))

        # Sheet 3: Delayed
        self._create_category_sheet(wb, "Delayed", sla_data.get("delayed", []))

        # Sheet 4: At-Risk (≤24h)
        self._create_category_sheet(wb, "At-Risk", sla_data.get("at_risk", []))

        # Sheet 5: Breached
        self._create_category_sheet(wb, "Breached", sla_data.get("breached", []))

        # Sheet 6: Critical SLA
        self._create_category_sheet(wb, "Critical SLA", sla_data.get("critical_sla", []))

        # Save workbook
        wb.save(output_path)

        return output_path

    def _create_summary_sheet(self, wb: Workbook, data: dict, sla_data: dict):
        """Create summary sheet with compliance overview."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "SLA Compliance Summary"
        ws["A1"].font = Font(size=16, bold=True, color="E74C3C")
        ws.merge_cells("A1:D1")

        # Metadata
        ws["A3"] = "Entity:"
        ws["B3"] = data["entity"]
        ws["A4"] = "Period:"
        ws["B4"] = data["period"]
        ws["A5"] = "Generated:"
        ws["B5"] = data["generated_at"].strftime("%Y-%m-%d %H:%M:%S")

        # Compliance Metrics
        ws["A7"] = "Compliance Metrics"
        ws["A7"].font = Font(size=14, bold=True)

        metrics = [
            ["Metric", "Count", "Percentage"],
            ["Total Accounts", sla_data["total_accounts"], "100%"],
            [
                "On-Time / Compliant",
                sla_data["compliant_count"],
                f"{sla_data['compliance_rate']:.1f}%",
            ],
            [
                "At-Risk (≤24h)",
                sla_data["at_risk_count"],
                f"{(sla_data['at_risk_count'] / sla_data['total_accounts'] * 100) if sla_data['total_accounts'] > 0 else 0:.1f}%",
            ],
            [
                "Breached",
                sla_data["breached_count"],
                f"{(sla_data['breached_count'] / sla_data['total_accounts'] * 100) if sla_data['total_accounts'] > 0 else 0:.1f}%",
            ],
        ]

        for row_idx, row_data in enumerate(metrics, start=8):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 8:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="E74C3C", end_color="E74C3C", fill_type="solid"
                    )
                    cell.font = Font(bold=True, color="FFFFFF")

        # Color code based on compliance rate
        compliance_rate = sla_data["compliance_rate"]
        ws["A14"] = "Overall Compliance Status:"
        if compliance_rate >= 95:
            ws["B14"] = "✓ Excellent"
            ws["B14"].font = Font(color="27AE60", bold=True)
        elif compliance_rate >= 80:
            ws["B14"] = "⚠ Good"
            ws["B14"].font = Font(color="F39C12", bold=True)
        else:
            ws["B14"] = "🔴 Needs Attention"
            ws["B14"].font = Font(color="E74C3C", bold=True)

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def _create_category_sheet(self, wb: Workbook, sheet_name: str, items: list[dict]):
        """Create a category-specific sheet (On-Time, At-Risk, etc.)."""
        ws = wb.create_sheet(sheet_name)

        ws["A1"] = f"{sheet_name} Items"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:G1")

        if items:
            # Headers
            headers = [
                "Account Code",
                "Account Name",
                "Status",
                "Criticality",
                "Balance",
                "Deadline",
                "Hours to Deadline",
                "Reviewer",
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)

                # Color code headers by category
                if sheet_name == "Breached":
                    cell.fill = PatternFill(
                        start_color="C0392B", end_color="C0392B", fill_type="solid"
                    )
                elif sheet_name == "At-Risk":
                    cell.fill = PatternFill(
                        start_color="E67E22", end_color="E67E22", fill_type="solid"
                    )
                elif sheet_name == "On-Time":
                    cell.fill = PatternFill(
                        start_color="27AE60", end_color="27AE60", fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color="3498DB", end_color="3498DB", fill_type="solid"
                    )

                cell.font = Font(bold=True, color="FFFFFF")

            # Data rows
            for row_idx, item in enumerate(items, start=4):
                ws.cell(row=row_idx, column=1, value=item["account_code"])
                ws.cell(row=row_idx, column=2, value=item["account_name"])
                ws.cell(row=row_idx, column=3, value=item["review_status"])
                ws.cell(row=row_idx, column=4, value=item["criticality"])
                ws.cell(row=row_idx, column=5, value=item["closing_balance"]).number_format = (
                    "₹#,##0"
                )
                ws.cell(row=row_idx, column=6, value=item["deadline"].strftime("%Y-%m-%d %H:%M"))
                ws.cell(row=row_idx, column=7, value=f"{item['time_to_deadline_hours']:.1f}")
                ws.cell(row=row_idx, column=8, value=item["reviewer"])

                # Color code time to deadline
                time_cell = ws.cell(row=row_idx, column=7)
                if item["time_to_deadline_hours"] < 0:
                    time_cell.font = Font(color="C0392B", bold=True)
                elif item["time_to_deadline_hours"] <= 24:
                    time_cell.font = Font(color="E67E22", bold=True)
        else:
            ws["A3"] = f"No {sheet_name.lower()} items."

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def _generate_pdf(self, data: dict) -> Path:
        """Generate PDF compliance summary."""
        output_path = self._get_output_path("pdf")

        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=20,
            textColor=colors.HexColor("#e74c3c"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        elements.append(Paragraph("SLA Compliance Report", title_style))
        elements.append(
            Paragraph(f"Entity: {data['entity']} | Period: {data['period']}", styles["Normal"])
        )
        elements.append(Spacer(1, 0.3 * inch))

        # Check for error
        if "error" in data:
            elements.append(Paragraph(f"Error: {data['error']}", styles["Normal"]))
            doc.build(elements)
            return output_path

        sla_data = data.get("sla_data", {})

        # Compliance Summary
        elements.append(Paragraph("Compliance Overview", styles["Heading2"]))

        summary_data = [
            ["Metric", "Count", "Percentage"],
            ["Total Accounts", str(sla_data["total_accounts"]), "100%"],
            [
                "On-Time / Compliant",
                str(sla_data["compliant_count"]),
                f"{sla_data['compliance_rate']:.1f}%",
            ],
            [
                "At-Risk (≤24h)",
                str(sla_data["at_risk_count"]),
                f"{(sla_data['at_risk_count'] / sla_data['total_accounts'] * 100) if sla_data['total_accounts'] > 0 else 0:.1f}%",
            ],
            [
                "Breached",
                str(sla_data["breached_count"]),
                f"{(sla_data['breached_count'] / sla_data['total_accounts'] * 100) if sla_data['total_accounts'] > 0 else 0:.1f}%",
            ],
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e74c3c")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 1), (2, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]
            )
        )

        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        # At-Risk Items
        elements.append(Paragraph("At-Risk Items (≤24h to Deadline)", styles["Heading2"]))

        at_risk_items = sla_data.get("at_risk", [])[:10]  # Top 10 most urgent

        if at_risk_items:
            risk_data = [["Account", "Criticality", "Hours Left", "Reviewer"]]
            for item in at_risk_items:
                risk_data.append(
                    [
                        f"{item['account_code']} - {item['account_name'][:20]}",
                        item["criticality"],
                        f"{item['time_to_deadline_hours']:.1f}",
                        item["reviewer"][:20],
                    ]
                )

            risk_table = Table(risk_data, colWidths=[2.5 * inch, 1.2 * inch, 1 * inch, 1.8 * inch])
            risk_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f39c12")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (2, 1), (2, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )

            elements.append(risk_table)
        else:
            elements.append(Paragraph("✓ No items at immediate risk.", styles["Normal"]))

        elements.append(Spacer(1, 0.3 * inch))

        # Breached Items
        elements.append(Paragraph("Breached SLA Items", styles["Heading2"]))

        breached_items = sla_data.get("breached", [])[:10]

        if breached_items:
            breach_data = [["Account", "Criticality", "Overdue (Hours)", "Reviewer"]]
            for item in breached_items:
                breach_data.append(
                    [
                        f"{item['account_code']} - {item['account_name'][:20]}",
                        item["criticality"],
                        f"{abs(item['time_to_deadline_hours']):.1f}",
                        item["reviewer"][:20],
                    ]
                )

            breach_table = Table(
                breach_data, colWidths=[2.5 * inch, 1.2 * inch, 1 * inch, 1.8 * inch]
            )
            breach_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#c0392b")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (2, 1), (2, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )

            elements.append(breach_table)
        else:
            elements.append(Paragraph("✓ No SLA breaches.", styles["Normal"]))

        elements.append(Spacer(1, 0.3 * inch))

        # Recommendations
        elements.append(Paragraph("Recommendations", styles["Heading2"]))

        recommendations = self._generate_recommendations(sla_data)

        rec_style = ParagraphStyle(
            "Recommendations", parent=styles["Normal"], leftIndent=20, spaceAfter=10
        )

        for rec in recommendations:
            elements.append(Paragraph(f"• {rec}", rec_style))

        # Footer
        elements.append(Spacer(1, 0.3 * inch))
        footer_style = ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER,
        )
        elements.append(
            Paragraph(
                f"Generated by Project Aura | {data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}",
                footer_style,
            )
        )

        # Build PDF
        doc.build(elements)

        return output_path

    def _generate_recommendations(self, sla_data: dict) -> list[str]:
        """Generate actionable SLA recommendations."""
        recommendations = []

        compliance_rate = sla_data["compliance_rate"]
        at_risk_count = sla_data["at_risk_count"]
        breached_count = sla_data["breached_count"]

        # Compliance rate recommendations
        if compliance_rate < 80:
            recommendations.append(
                f"CRITICAL: SLA compliance is {compliance_rate:.1f}%, below acceptable threshold. "
                "Immediate action required to improve review turnaround times."
            )
        elif compliance_rate < 95:
            recommendations.append(
                f"SLA compliance is {compliance_rate:.1f}%. Focus on reducing at-risk and breached items."
            )
        else:
            recommendations.append(
                f"SLA compliance is excellent at {compliance_rate:.1f}%. Maintain current processes."
            )

        # At-risk items
        if at_risk_count > 0:
            recommendations.append(
                f"Prioritize {at_risk_count} at-risk items (≤24h to deadline) to prevent SLA breaches. "
                "Reassign or escalate as needed."
            )

        # Breached items
        if breached_count > 0:
            recommendations.append(
                f"Address {breached_count} breached SLA items immediately. "
                "Investigate root causes and implement corrective actions."
            )

        # Critical SLA tracking
        critical_count = len(sla_data.get("critical_sla", []))
        if critical_count > 0:
            recommendations.append(
                f"Monitor {critical_count} critical SLA items closely. "
                "These require daily tracking to ensure timely completion."
            )

        return recommendations
