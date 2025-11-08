"""
Variance Analysis Report Generator

Generates period-over-period variance analysis with waterfall charts, heatmaps,
and detailed variance explanations in PDF and Excel formats.
"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from src.analytics import calculate_variance_analysis, perform_analytics
from src.reports import BaseReport, register_report
from src.visualizations import create_variance_waterfall_chart, export_chart_to_html


@register_report("variance")
class VarianceAnalysisReport(BaseReport):
    """Generate period-over-period variance analysis report."""

    def generate(self) -> dict[str, str]:
        """
        Generate variance report in PDF and Excel formats.

        Returns:
            Dict with 'pdf', 'excel', and 'html' keys mapping to file paths
        """
        self.generated_at = datetime.now()

        # Get previous period (default: one month back)
        previous_period = self.kwargs.get("previous_period", self._get_previous_period())

        # Fetch data
        data = self._fetch_data(previous_period)

        # Generate Excel (primary format with charts)
        excel_path = self._generate_excel(data, previous_period)
        self.generated_files["excel"] = str(excel_path)

        # Generate PDF summary
        pdf_path = self._generate_pdf(data, previous_period)
        self.generated_files["pdf"] = str(pdf_path)

        # Generate HTML chart
        if "variance_analysis" in data and not data["variance_analysis"].get("error"):
            html_path = self._generate_html_chart(data)
            self.generated_files["html"] = str(html_path)

        return self.generated_files

    def _get_previous_period(self) -> str:
        """Calculate previous period from current period."""
        try:
            year, month = map(int, self.period.split("-"))
            if month == 1:
                prev_month = 12
                prev_year = year - 1
            else:
                prev_month = month - 1
                prev_year = year
            return f"{prev_year:04d}-{prev_month:02d}"
        except:
            return self.period  # Fallback to same period

    def _fetch_data(self, previous_period: str) -> dict:
        """Fetch variance and analytics data."""
        try:
            # Get variance analysis
            variance_analysis = calculate_variance_analysis(
                self.entity, self.period, previous_period
            )

            # Get current period analytics
            current_analytics = perform_analytics(self.entity, self.period)

            # Get previous period analytics
            previous_analytics = perform_analytics(self.entity, previous_period)

            return {
                "variance_analysis": variance_analysis,
                "current_analytics": current_analytics,
                "previous_analytics": previous_analytics,
                "entity": self.entity,
                "current_period": self.period,
                "previous_period": previous_period,
                "generated_at": self.generated_at,
            }
        except Exception as e:
            return {
                "error": str(e),
                "entity": self.entity,
                "current_period": self.period,
                "previous_period": previous_period,
                "generated_at": self.generated_at,
            }

    def _generate_excel(self, data: dict, previous_period: str) -> Path:
        """Generate multi-sheet Excel workbook."""
        output_path = self._get_output_path("xlsx")

        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Sheet 1: Summary
        self._create_summary_sheet(wb, data, previous_period)

        # Sheet 2: Top Movers
        self._create_top_movers_sheet(wb, data)

        # Sheet 3: Variance Details
        self._create_variance_details_sheet(wb, data)

        # Sheet 4: Category Analysis
        self._create_category_analysis_sheet(wb, data)

        # Sheet 5: Trend Analysis
        self._create_trend_analysis_sheet(wb, data)

        # Sheet 6: Charts (placeholder for embedded charts)
        self._create_charts_sheet(wb, data)

        # Save workbook
        wb.save(output_path)

        return output_path

    def _create_summary_sheet(self, wb: Workbook, data: dict, previous_period: str):
        """Create summary sheet with key metrics."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "Variance Analysis Summary"
        ws["A1"].font = Font(size=16, bold=True, color="1F77B4")
        ws.merge_cells("A1:D1")

        # Metadata
        ws["A3"] = "Entity:"
        ws["B3"] = data["entity"]
        ws["A4"] = "Current Period:"
        ws["B4"] = data["current_period"]
        ws["A5"] = "Previous Period:"
        ws["B5"] = data["previous_period"]
        ws["A6"] = "Generated:"
        ws["B6"] = data["generated_at"].strftime("%Y-%m-%d %H:%M:%S")

        # Key Metrics
        ws["A8"] = "Key Metrics"
        ws["A8"].font = Font(size=14, bold=True)

        variance_data = data.get("variance_analysis", {})
        current_data = data.get("current_analytics", {})
        previous_data = data.get("previous_analytics", {})

        metrics = [
            ["Metric", "Current Period", "Previous Period", "Variance", "Variance %"],
            [
                "Total Balance",
                current_data.get("total_balance", 0),
                previous_data.get("total_balance", 0),
                variance_data.get("total_variance", 0),
                (
                    f"{(variance_data.get('total_variance', 0) / previous_data.get('total_balance', 1) * 100):.2f}%"
                    if previous_data.get("total_balance", 0) != 0
                    else "N/A"
                ),
            ],
            [
                "Accounts Analyzed",
                current_data.get("account_count", 0),
                variance_data.get("accounts_analyzed", 0),
                0,
                "0%",
            ],
            ["Significant Variances", variance_data.get("significant_count", 0), "-", "-", "-"],
        ]

        for row_idx, row_data in enumerate(metrics, start=9):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 9:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="3498DB", end_color="3498DB", fill_type="solid"
                    )
                    cell.font = Font(bold=True, color="FFFFFF")
                if col_idx >= 2 and row_idx > 9 and isinstance(value, (int, float)):
                    cell.number_format = "₹#,##0"

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_top_movers_sheet(self, wb: Workbook, data: dict):
        """Create top movers sheet showing largest variances."""
        ws = wb.create_sheet("Top Movers")

        ws["A1"] = "Top Movers - Significant Variances"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:F1")

        variance_data = data.get("variance_analysis", {})
        significant_variances = variance_data.get("significant_variances", [])

        if significant_variances:
            # Headers
            headers = [
                "Account Code",
                "Account Name",
                "Current Balance",
                "Previous Balance",
                "Variance",
                "Variance %",
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Data rows
            for row_idx, var in enumerate(significant_variances, start=4):
                ws.cell(row=row_idx, column=1, value=var.get("account_code", ""))
                ws.cell(row=row_idx, column=2, value=var.get("account_name", ""))
                ws.cell(
                    row=row_idx, column=3, value=var.get("current_balance", 0)
                ).number_format = "₹#,##0"
                ws.cell(
                    row=row_idx, column=4, value=var.get("previous_balance", 0)
                ).number_format = "₹#,##0"
                ws.cell(row=row_idx, column=5, value=var.get("variance", 0)).number_format = (
                    "₹#,##0"
                )
                ws.cell(row=row_idx, column=6, value=f"{var.get('variance_pct', 0):.2f}%")

                # Color code variance
                variance_cell = ws.cell(row=row_idx, column=5)
                if var.get("variance", 0) > 0:
                    variance_cell.font = Font(color="27AE60")  # Green
                elif var.get("variance", 0) < 0:
                    variance_cell.font = Font(color="E74C3C")  # Red
        else:
            ws["A3"] = "No significant variances detected."

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def _create_variance_details_sheet(self, wb: Workbook, data: dict):
        """Create detailed variance breakdown sheet."""
        ws = wb.create_sheet("Variance Details")

        ws["A1"] = "Detailed Variance Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        variance_data = data.get("variance_analysis", {})

        # Summary stats
        ws["A3"] = "Accounts Analyzed:"
        ws["B3"] = variance_data.get("accounts_analyzed", 0)
        ws["A4"] = "Total Variance:"
        ws["B4"] = variance_data.get("total_variance", 0)
        ws["B4"].number_format = "₹#,##0"
        ws["A5"] = "Significant Variances:"
        ws["B5"] = variance_data.get("significant_count", 0)

        # All variances table (if available in data)
        significant_variances = variance_data.get("significant_variances", [])
        if significant_variances:
            ws["A7"] = "All Significant Variances"
            ws["A7"].font = Font(bold=True)

            # Convert to DataFrame for easier manipulation
            df = pd.DataFrame(significant_variances)

            # Write DataFrame to sheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=8):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

    def _create_category_analysis_sheet(self, wb: Workbook, data: dict):
        """Create category-wise variance analysis."""
        ws = wb.create_sheet("Category Analysis")

        ws["A1"] = "Variance by Category"
        ws["A1"].font = Font(size=14, bold=True)

        current_data = data.get("current_analytics", {})
        previous_data = data.get("previous_analytics", {})

        current_by_cat = current_data.get("by_category", {})
        previous_by_cat = previous_data.get("by_category", {})

        # Headers
        headers = ["Category", "Current Balance", "Previous Balance", "Variance", "Variance %"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data rows
        all_categories = set(list(current_by_cat.keys()) + list(previous_by_cat.keys()))
        for row_idx, category in enumerate(sorted(all_categories), start=4):
            current_bal = current_by_cat.get(category, 0)
            previous_bal = previous_by_cat.get(category, 0)
            variance = current_bal - previous_bal
            variance_pct = (variance / previous_bal * 100) if previous_bal != 0 else 0

            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=current_bal).number_format = "₹#,##0"
            ws.cell(row=row_idx, column=3, value=previous_bal).number_format = "₹#,##0"
            ws.cell(row=row_idx, column=4, value=variance).number_format = "₹#,##0"
            ws.cell(row=row_idx, column=5, value=f"{variance_pct:.2f}%")

    def _create_trend_analysis_sheet(self, wb: Workbook, data: dict):
        """Create trend analysis sheet."""
        ws = wb.create_sheet("Trend Analysis")

        ws["A1"] = "Period-over-Period Trends"
        ws["A1"].font = Font(size=14, bold=True)

        ws["A3"] = "Two-period trend analysis"
        ws["A4"] = "For multi-period trends, use Multi-Period Comparison feature in Analytics"

    def _create_charts_sheet(self, wb: Workbook, data: dict):
        """Create charts sheet (placeholder for embedded charts)."""
        ws = wb.create_sheet("Charts")

        ws["A1"] = "Variance Charts"
        ws["A1"].font = Font(size=14, bold=True)

        ws["A3"] = "Waterfall Chart: See HTML export for interactive waterfall visualization"
        ws["A4"] = "Heatmap: Available in dashboard view"
        ws["A5"] = "Trend Lines: See Trend Analysis sheet"

    def _generate_pdf(self, data: dict, previous_period: str) -> Path:
        """Generate PDF summary report."""
        output_path = self._get_output_path("pdf")

        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=20,
            textColor=colors.HexColor("#2c3e50"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        elements.append(Paragraph("Variance Analysis Report", title_style))
        elements.append(
            Paragraph(
                f"Entity: {data['entity']} | Current: {data['current_period']} | "
                f"Previous: {data['previous_period']}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 0.3 * inch))

        # Summary Table
        variance_data = data.get("variance_analysis", {})

        summary_data = [
            ["Metric", "Value"],
            ["Accounts Analyzed", str(variance_data.get("accounts_analyzed", 0))],
            ["Total Variance", f"₹{variance_data.get('total_variance', 0):,.0f}"],
            ["Significant Variances", str(variance_data.get("significant_count", 0))],
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]
            )
        )

        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Top Movers
        elements.append(Paragraph("Top Variance Drivers", styles["Heading2"]))

        significant_variances = variance_data.get("significant_variances", [])[:10]
        if significant_variances:
            var_data = [["Account", "Variance", "Variance %"]]
            for var in significant_variances:
                var_data.append(
                    [
                        f"{var.get('account_code')} - {var.get('account_name', '')[:20]}",
                        f"₹{var.get('variance', 0):,.0f}",
                        f"{var.get('variance_pct', 0):.1f}%",
                    ]
                )

            var_table = Table(var_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
            var_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e74c3c")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )

            elements.append(var_table)
        else:
            elements.append(Paragraph("No significant variances detected.", styles["Normal"]))

        # Build PDF
        doc.build(elements)

        return output_path

    def _generate_html_chart(self, data: dict) -> Path:
        """Generate interactive HTML waterfall chart."""
        variance_data = data.get("variance_analysis", {})
        significant_variances = variance_data.get("significant_variances", [])

        if significant_variances:
            df = pd.DataFrame(significant_variances)
            fig = create_variance_waterfall_chart(
                df, title=f"Variance Analysis: {self.entity} - {self.period}"
            )

            output_path = self._get_output_path("html", suffix="waterfall")
            export_chart_to_html(fig, str(output_path))

            return output_path

        return None
